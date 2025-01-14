import socket
import pyrealsense2 as rs
import cv2
import base64
import json
import numpy as np
import time 
import threading
import struct
from datetime import datetime 
import openpyxl
import os
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import traceback

sheetNames = {
        'MS-010' : 'Aurora1 8 SUS304 Male tube 성형',
        'MS-011' : 'Aurora1 12 SUS304  Male tube 성형',
        'MS-012' : 'Aurora1 8 SUS304 이중축관성형',
        'MS-013' : 'Aurora1 12 SUS304 이중축관성형',
        'MS-014' : 'Aurora1 4.76사두 성형',
        'MS-015' : 'Aurora1 4.76나팔 120° FLARE성형'
    }

cellIndex = {
    'date': ['L7', 'T7', 'AB7'],
    'values': ['L10',
                'T10',
                'AB10',
                ]
}

json_data = None
plc_lock = threading.Lock()
flutter_lock = threading.Lock()

global data
data = None
input_flag = True

names = ['MS-010', 'MS-011', 'MS-012', 'MS-013', 'MS-014', 'MS-015']
MS010 = [
    'MS-010',
    [10.98],
    ["3"],
    [0.15, -0.15],
    {
      "2": [7.89, -0.06, 0.06],
      "4": [4.95, -0.25, 0.25],
      "5": [1.7, -0.1, 0.1],
      "6": [21.12, -0.25, 0.25]
    },
    13
  ]
MS011 = [
    'MS-011',
    [16.51],
    ["3"],
    [0.25, -0.25],
    {
      "2": [11.8, -0.1, 0.1],
      "4": [7.75, -0.25, 0.25],
      "5": [2.54, -0.2, 0.2],
      "6": [26.62, -0.5, 0.5]
    },
    13
  ]
MS012 = [
    'MS-012',
    [8.8, 9.0],
    ["3", "4"],
    [0.2, -0.2, 0.2, -0.2],
    {
      "2": [6.75, -0.1, 0.1],
      "5": [1.4, -0.1, 0.1]
    },
    13
  ]
MS013 = [
    'MS-013',
    [13.5],
    ["3"],
    [0.2, -0.2],
    {
      "2": [11.45, -0.1, 0.1],
      "4": [13.5, -0.2, 0.2],
      "5": [1.4, -0.1, 0.1]
    },
    13
  ]
MS014 = [
    'MS-014',
    [7.1, 3.2, 115.0],
    ["2", "3", "4"],
    [0.4, -0.18, 0.2, -0.1, 2.0, -2.0],
    [0.38, -0.16, 0.18, -0.08, 1.97, -1.97],
    {
      "5": [3.5, -0.5, 0.5, -0.47, 0.47],
      "6": [5.11, -0.08, 0.08, -0.07, 0.07],
      "7": [0, -0.2, 0.2, -0.19, 0.19]
    },
    14
  ]
MS015 = [
    'MS-015',
    [7.1, 3.2, 120.0],
    ["2", "3", "5"],
    [0.4, -0.2, 0.2, -0.1, 2.0, -2.0],
    [0.38, -0.18, 0.18, -0.08, 1.97, -1.97],
    {
      "4": [1.4, -0.2, 0.2, -0.19, 0.19],
      "6": [3.5, -0.5, 0.5, -0.47, 0.47],
      "7": [5.11, -0.08, 0.08, -0.07, 0.07],
      "8": [0, -0.2, 0.2, -0.19, 0.19]
    },
    15
  ]

dataPath = "C:\\UND\\data.json";
dateListPath = "C:\\UND\\dateList.json";
excelPath = "C:\\UND\\form.xlsx";

class Server:
    def __init__(self, flutter_host, flutter_port, plc_ip, plc_port, plc_flag_addr, plc_flag_length, plc_data_addr, plc_data_length):
        self.flutter_host = flutter_host
        self.flutter_port = flutter_port
        self.flutter_socket = None
        self.is_flutter_connected = False
        self.flutter_received = False
         
        self.plc_ip = plc_ip
        self.plc_port = plc_port
        self.plc_flag_addr = plc_flag_addr
        self.plc_flag_length = plc_flag_length
        self.plc_data_addr = plc_data_addr
        self.plc_data_length = plc_data_length
        self.plc_received = False
        self.plc_connection = None

        # 쓰레드 공유 데이터
        self.plc_data = None
        
        self.pipeline = rs.pipeline()
        config = rs.config()
        config.enable_stream(rs.stream.color, 640, 480, rs.format.bgr8, 30)  # 컬러 영상 설정
        self.pipeline.start(config)
        self.data = None


    def set_plc_data(self, data):
        plc_lock.acquire()
        self.plc_data = data
        self.plc_received = True
        plc_lock.release()

    def init_plc_data(self, data):
        plc_lock.acquire()
        self.plc_data = None
        self.plc_received = False
        plc_lock.release()

    def combine_registers(response):
        high_reg = response.registers[0]
        low_reg = response.registers[1]
        
        combined = (high_reg << 16) | low_reg

        little_endian_values = [
            int.from_bytes(value.to_bytes(2, byteorder='big'), byteorder='little')
            for value in response.registers
        ]
        big_endian_values = [
            int.from_bytes(value.to_bytes(2, byteorder='little'), byteorder='big')
            for value in response.registers
        ]
        return struct.unpack('>f', combined.to_bytes(4, byteorder='big'))[0]
 
    def connect_plc(self):
        # 연결 끊김 시 재 연결 시도
        while True:
            try:
                client = ModbusTcpClient(self.plc_ip, port=self.plc_port)
                self.plc_connection = True

                while True:
                    if client.connect():
                        
                        while True:
                            if self.plc_received:
                                time.sleep(1)
                                continue
                            
                            response = client.read_holding_registers(self.plc_flag_addr, count=self.plc_flag_length)
                            result = self.combine_registers(response)
                            if result != 0:
                                break
                            
                            response = client.read_holding_registers(self.plc_data_addr, count=self.plc_data_length)
                            result = self.combine_registers(response)
                        
                            self.set_plc_data(result)
                            
                        time.sleep(5)
                        
                    else:
                        print("Unable to connect to Modbus server")
                        break
            except Exception as e:
                print(f'error : {e}')
                pass
            finally:
                self.plc_connection = False
                client.close()
    
    def connect_plc1(self):
        """PLC와 EtherNet/IP 통신 연결"""
        try:
            print(f"Connecting to PLC at {self.plc_ip}...")
            self.plc_connection = LogixDriver(self.plc_ip)
            print("PLC connected successfully.")
        except Exception as e:
            print(f"Failed to connect to PLC: {e}")
            raise

    def create_data(self):
        return None

    def input_func(self):
        global data
        
        while input_flag:
            time.sleep(1)
            
            date = input("date: ")
            name = input("name: ")
            values = []
            
            while True:
                string_value = input("input values :")
                if string_value == 'q':
                    break
                values.append(float(string_value))

            # 메시지와 함께 JSON 데이터 생성
            self.data = {
                "date": date,
                "name": name,
                "values" : values,
            }

    def read_and_create_xl(self, conn):
        while True:
            try:
                # 데이터 수신
                while True:
                    received_data = conn.recv(4096).decode('utf-8')
                    if received_data:
                        # 연결이 종료된 경우 루프를 나감
                        print("엑셀 데이터 수신")
                        break
                # JSON 문자열을 Python 딕셔너리로 변환
                data = json.loads(received_data)

                # 변수로 저장
                self.save_excel_file(data, excelPath)
                
                print('엑셀 파일 저장 성공')
                print('date : ')
                
            except Exception as e:
                print('read_and_create_xl ')
                print(e)

    
    def send_data(self, conn):
        try:
            # pipeline = rs.pipeline()
            # config = rs.config()
            # config.enable_stream(rs.stream.color, 640, 480, rs.format.bgr8, 30)  # 컬러 영상 설정
            # pipeline.start(config)

            while True:
                
                # RealSense 카메라에서 프레임 캡처
                frames = self.pipeline.wait_for_frames()
                color_frame = frames.get_color_frame()
                if not color_frame:
                    continue
                
                # 이미지로 변환
                frame = np.asanyarray(color_frame.get_data())

                # 실시간 비디오 표시
                # cv2.imshow("Real-Time Video", frame)

                # 이미지 Base64로 인코딩
                _, buffer = cv2.imencode('.jpg', frame)
                image_base64 = base64.b64encode(buffer).decode('utf-8')
                
                if self.data != None:
                # 메시지와 함께 JSON 데이터 생성
                    self.data["image"] = image_base64
                    json_data = json.dumps(self.data)

                    # JSON 데이터 전송 (길이 정보 포함)
                    
                    conn.sendall(len(json_data).to_bytes(4, 'big') + json_data.encode('utf-8'))
                    print("Data sent!")
                    self.data = None
                else:
                    continue     
            
                time.sleep(10)
        except Exception as e:
            print('send_data ')
            print({e})
            

    def flutter_test(self):
        global data
        while True:
            try:
                flutter_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                flutter_socket.bind((self.flutter_host, self.flutter_port))
                flutter_socket.listen(1)
                print("Waiting for connection...")

                conn, addr = flutter_socket.accept()
                print(f"Connected by {addr}")
                input_flag = True
                
                input_thread = threading.Thread(target=self.input_func)
                input_thread.start()
                
                read_thread = threading.Thread(target=self.send_data, args=(conn,))
                read_thread.start()
                
                self.read_and_create_xl(conn)                
                
                # # RealSense 카메라 파이프라인 설정
        
            except Exception as e:
                print(f"Terminating... {e}")
            finally:
                input_flag = False
                conn.close()
                flutter_socket.close()


                
    def save_excel_file(self, map_data, excel_path = dataPath):
        first_date = ''
        
        try:
            # 1. 기존 Excel 파일을 읽기
            if not os.path.exists(excel_path):
                raise Exception(f"파일이 존재하지 않습니다: {excel_path}")

            workbook = openpyxl.load_workbook(excel_path)
            
            date_index = 0
            data_index = 0

            for name in map_data.keys():
                sheet_name = sheetNames[name]
                sheet = workbook[sheet_name]

                date_list = map_data[name].keys()
                sorted_date_list = sorted(date_list, key=lambda date: datetime.strptime(date, "%Y-%m-%d"))
                
                for date in sorted_date_list:
                    if first_date == '':
                        first_date = date

                    data = map_data[name][date] if date in map_data[name] else None
                    if data is None:
                        continue

                    # Date Cell 위치 계산
                    date_cell_position = cellIndex['date'][date_index % 3]
                    date_cell_coordinates = _parse_cell_position(date_cell_position)

                    if not date_cell_coordinates:
                        raise Exception(f"잘못된 셀 위치 형식입니다: {date_cell_position}")
 
                    tmp_list = getDataParams(name)
                    
                    row, col = date_cell_coordinates
                    if date_index > 2:
                        row += tmp_list[5]
                    sheet.cell(row=row + 1, column=col + 1).value = date

                    
                    for check_num in data['measurements'].keys():
                        for value in data['measurements'][check_num]:
                            
                            value_cell_position = cellIndex['values'][date_index % 3]
                            value_cell_coordinates = _parse_cell_position(value_cell_position)

                            if not value_cell_coordinates:
                                raise Exception(f"잘못된 셀 위치 형식입니다: {value_cell_position}")
                            
                            row = value_cell_coordinates[0] + int(check_num) - 1
                            
                            if date_index > 2:
                                row += tmp_list[5]

                            col = (value_cell_coordinates[1] + 2 * data_index if data_index != len(data['measurements'][check_num]) - 1
                                else value_cell_coordinates[1] + 2 * 3)

                            sheet.cell(row=row + 1, column=col + 1).value = value

                            # new_file_path = f"C:\\UND\\{first_date}.xlsx"
                            # workbook.save(new_file_path)
                            # print(f"파일이 성공적으로 저장되었습니다: {first_date}")

                            data_index +=1
                        data_index = 0
                    date_index += 1

                date_index = 0

            # 4. 새 파일로 저장
            new_file_path = f"C:\\UND\\{first_date}.xlsx"
            workbook.save(new_file_path)
            print(f"파일이 성공적으로 저장되었습니다: {first_date}")

        except Exception as e:
            print(f"오류 발생: {traceback.format_exc()}")


    def getDataParams(self, name) :
        match  name: 
            case 'MS-010':
                return MS010
            case 'MS-011':
                return MS011
            case 'MS-012':
                return MS012
            case 'MS-013':
                return MS013
            case 'MS-014':
                return MS014
            case 'MS-015':
                return MS015
        return []
  
    def _parse_cell_position(position):
        """셀 위치 문자열(A1 형식)을 (행, 열) 튜플로 변환"""
        try:
            coord = coordinate_from_string(position)
            col = column_index_from_string(coord[0]) - 1  # 열 번호는 0부터 시작
            row = coord[1] - 1  # 행 번호는 0부터 시작
            return (row, col)
        except Exception as e:
            print(f"셀 위치 파싱 오류: {e}")
            return None



def flutter_test(host, port):
    global data
    while True:
        try:
            flutter_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            flutter_socket.bind((host, port))
            flutter_socket.listen(1)
            print("Waiting for connection...")

            conn, addr = flutter_socket.accept()
            print(f"Connected by {addr}")
            input_flag = True
            
            input_thread = threading.Thread(target=input_func)
            input_thread.start()
            
            # # RealSense 카메라 파이프라인 설정
    
        except Exception as e:
            print(f"Terminating... {e}")
        finally:
            input_flag = False
            conn.close()
            flutter_socket.close()
            
            
def send_data(self, conn):
    try:
        # pipeline = rs.pipeline()
        # config = rs.config()
        # config.enable_stream(rs.stream.color, 640, 480, rs.format.bgr8, 30)  # 컬러 영상 설정
        # pipeline.start(config)

        while True:
            
            # RealSense 카메라에서 프레임 캡처
            frames = self.pipeline.wait_for_frames()
            color_frame = frames.get_color_frame()
            if not color_frame:
                continue
            
            # 이미지로 변환
            frame = np.asanyarray(color_frame.get_data())

            # 실시간 비디오 표시
            # cv2.imshow("Real-Time Video", frame)

            # 이미지 Base64로 인코딩
            _, buffer = cv2.imencode('.jpg', frame)
            image_base64 = base64.b64encode(buffer).decode('utf-8')
            
            if data != None:
            # 메시지와 함께 JSON 데이터 생성
                data["image"] = image_base64
                json_data = json.dumps(data)

                # JSON 데이터 전송 (길이 정보 포함)
                
                conn.sendall(len(json_data).to_bytes(4, 'big') + json_data.encode('utf-8'))
                print("Data sent!")
                data = None
            else:
                continue     
        
            time.sleep(10)
    except Exception as e:
        print({e})
        
    finally:
        self.pipeline.stop()  # RealSense 카메라 종료
        cv2.destroyAllWindows()  # OpenCV 창 닫기
            
# def save_excel_file(map_data, excel_path):
#     first_date = ''
    
#     try:
#         # 1. 기존 Excel 파일을 읽기
#         if not os.path.exists(excel_path):
#             raise Exception(f"파일이 존재하지 않습니다: {excel_path}")

#         workbook = openpyxl.load_workbook(excel_path)
        
#         date_index = 0
#         data_index = 0

#         for name in map_data.keys():
#             sheet_name = sheetNames[name]
#             sheet = workbook[sheet_name]

#             for date in map_data[name].keys():
#                 if first_date == '':
#                     first_date = date

#                 data = map_data[name][date] if date in map_data[name] else None
#                 if data is None:
#                     continue

#                 # Date Cell 위치 계산
#                 date_cell_position = cellIndex['date'][date_index]
#                 date_cell_coordinates = _parse_cell_position(date_cell_position)

#                 if not date_cell_coordinates:
#                     raise Exception(f"잘못된 셀 위치 형식입니다: {date_cell_position}")

#                 row, col = date_cell_coordinates
#                 sheet.cell(row=row + 1, column=col + 1).value = date

#                 for check_num in data['measurements'].keys():
#                     for i, measurement in enumerate(data['measurements'][check_num]):

#                         value_cell_position = cellIndex['values'][date_index]
#                         value_cell_coordinates = _parse_cell_position(value_cell_position)

#                         if not value_cell_coordinates:
#                             raise Exception(f"잘못된 셀 위치 형식입니다: {value_cell_position}")
#                         tmp_list = getDataParams(name)
#                         if date_index > 2:
#                             row = value_cell_coordinates[0] + tmp_list[5]
#                         else:
#                             row = value_cell_coordinates[0]

#                         col = (value_cell_coordinates[1] + 2 * data_index if i != len(data['measurements'][check_num]) - 1
#                                else value_cell_coordinates[1] + 2 * 3)

#                         sheet.cell(row=row + 1, column=col + 1).value = measurement


#                     data_index = 0
#                 date_index += 1

#             date_index = 0
            
#         # 4. 새 파일로 저장
#         new_file_path = f"C:\\UND\\{first_date}{data_index}.xlsx"
#         workbook.save(new_file_path)
#         print(f"파일이 성공적으로 저장되었습니다: {first_date}")

#     except Exception as e:
#         print(f"오류 발생: {e}")


def getDataParams(name) :
    match  name: 
      case 'MS-010':
        return MS010
      case 'MS-011':
        return MS011
      case 'MS-012':
        return MS012
      case 'MS-013':
        return MS013
      case 'MS-014':
        return MS014
      case 'MS-015':
        return MS015
    return []
  
def _parse_cell_position(position):
    """셀 위치 문자열(A1 형식)을 (행, 열) 튜플로 변환"""
    try:
        coord = coordinate_from_string(position)
        col = column_index_from_string(coord[0]) - 1  # 열 번호는 0부터 시작
        row = coord[1] - 1  # 행 번호는 0부터 시작
        return (row, col)
    except Exception as e:
        print(f"셀 위치 파싱 오류: {e}")
        return None

def input_func(self):
    global data
    
    while input_flag:
        time.sleep(1)
        
        date = input("date: ")
        name = input("name: ")
        values = []
        
        while True:
            string_value = input("input values :")
            if string_value == 'q':
                break
            values.append(float(string_value))
        
        # 메시지와 함께 JSON 데이터 생성
        data = {
            "date": date,
            "name": name,
            "values" : values,
        }

def main():
    flutter_host = "127.0.0.1"
    flutter_port = 12345
    plc_host = "127.0.0.1"
    plc_port = 12344
    
    plc_flag_addr = 5000
    plc_flag_length = 2
    plc_data_addr = 5002
    plc_data_length = 12
    
    
    # flutter_test(flutter_host, flutter_port)
    
    server = Server(flutter_host, flutter_port, plc_host, plc_port, 
                    plc_flag_addr, plc_flag_length, plc_data_addr, plc_data_length)

    server.flutter_test()

    return

if __name__ == "__main__":
    main()